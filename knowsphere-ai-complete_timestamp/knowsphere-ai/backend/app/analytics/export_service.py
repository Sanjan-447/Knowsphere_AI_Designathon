"""
Shared export utility — one function per format, all taking the same
(fieldnames, rows) shape, so any report can be exported as CSV, Excel, or
PDF without three separate implementations per report type.
"""
from __future__ import annotations

import csv
import io


def to_csv_bytes(fieldnames: list[str], rows: list[dict]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8")


def to_excel_bytes(fieldnames: list[str], rows: list[dict], sheet_name: str = "Report") -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name length limit

    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(f, "") for f in fieldnames])

    for col_idx, field in enumerate(fieldnames, start=1):
        max_len = max([len(str(field))] + [len(str(r.get(field, ""))) for r in rows]) if rows else len(str(field))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def to_pdf_bytes(title: str, fieldnames: list[str], rows: list[dict], subtitle: str = "") -> bytes:
    """A simple tabular PDF report — readable, not fancy. For large
    reports with many columns, CSV/Excel are the better format anyway;
    PDF here is meant for shareable summaries, not raw data dumps."""
    from fpdf import FPDF

    pdf = FPDF(orientation="L" if len(fieldnames) > 5 else "P")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, title, new_x="LMARGIN", new_y="NEXT")
    if subtitle:
        pdf.set_font("Helvetica", "", 9)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 6, subtitle, new_x="LMARGIN", new_y="NEXT")
        pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    if not rows:
        pdf.set_font("Helvetica", "I", 10)
        pdf.cell(0, 8, "No data for this report.", new_x="LMARGIN", new_y="NEXT")
        return bytes(pdf.output())

    usable_width = pdf.w - 2 * pdf.l_margin
    col_width = usable_width / len(fieldnames)

    pdf.set_font("Helvetica", "B", 8)
    pdf.set_fill_color(230, 230, 230)
    for field in fieldnames:
        pdf.cell(col_width, 7, str(field)[:40], border=1, fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    for row in rows[:500]:  # sane ceiling for a PDF table
        for field in fieldnames:
            value = str(row.get(field, ""))[:40]
            pdf.cell(col_width, 6, value, border=1)
        pdf.ln()

    return bytes(pdf.output())


def export_response(fmt: str, title: str, fieldnames: list[str], rows: list[dict], filename_prefix: str, subtitle: str = ""):
    """Builds a Flask Response for the requested format. Raises ValueError
    for an unsupported format, so callers can turn that into a 422."""
    from flask import Response

    if fmt == "csv":
        return Response(
            to_csv_bytes(fieldnames, rows), mimetype="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.csv"},
        )
    if fmt == "excel":
        return Response(
            to_excel_bytes(fieldnames, rows, sheet_name=title),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.xlsx"},
        )
    if fmt == "pdf":
        return Response(
            to_pdf_bytes(title, fieldnames, rows, subtitle=subtitle), mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename_prefix}.pdf"},
        )
    raise ValueError(f"Unsupported export format: {fmt}")
